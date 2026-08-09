from dataclasses import dataclass
from datetime import (
    date,
    datetime,
)
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

XLSX_ZIP_SIGNATURE = (
    b"PK\x03\x04"
)


@dataclass(
    frozen=True,
    slots=True,
)
class WorksheetInspection:
    title: str

    max_row: int
    max_column: int

    non_empty_rows: tuple[
        tuple[
            int,
            tuple[
                Any,
                ...
            ],
        ],
        ...
    ]


@dataclass(
    frozen=True,
    slots=True,
)
class WorkbookInspection:
    sheet_names: tuple[
        str,
        ...
    ]

    worksheets: tuple[
        WorksheetInspection,
        ...
    ]


def is_xlsx_bytes(
    content: bytes,
) -> bool:
    return content.startswith(
        XLSX_ZIP_SIGNATURE
    )


def normalize_cell_value(
    value: Any,
) -> Any:
    if isinstance(
        value,
        datetime,
    ):
        return value.isoformat()

    if isinstance(
        value,
        date,
    ):
        return value.isoformat()

    if isinstance(
        value,
        str,
    ):
        normalized = (
            " ".join(
                value.split()
            )
        )

        return (
            normalized
            if normalized
            else None
        )

    return value


def row_has_value(
    values: tuple[
        Any,
        ...
    ],
) -> bool:
    return any(
        value is not None
        and value != ""
        for value in values
    )


def inspect_xlsx_bytes(
    content: bytes,
    *,
    max_non_empty_rows: int = 40,
) -> WorkbookInspection:
    if not isinstance(
        content,
        bytes,
    ):
        raise TypeError(
            "XLSX content must be bytes."
        )

    if not is_xlsx_bytes(
        content
    ):
        raise ValueError(
            "Content does not have "
            "an XLSX ZIP signature."
        )

    if max_non_empty_rows <= 0:
        raise ValueError(
            "max_non_empty_rows "
            "must be positive."
        )

    workbook = load_workbook(
        filename=BytesIO(
            content
        ),
        read_only=True,
        data_only=True,
    )

    try:
        inspections = []

        for worksheet in (
            workbook.worksheets
        ):
            non_empty_rows = []

            for row_number, row in enumerate(
                worksheet.iter_rows(
                    values_only=True
                ),
                start=1,
            ):
                normalized = tuple(
                    normalize_cell_value(
                        value
                    )
                    for value in row
                )

                if not row_has_value(
                    normalized
                ):
                    continue

                non_empty_rows.append(
                    (
                        row_number,
                        normalized,
                    )
                )

                if (
                    len(
                        non_empty_rows
                    )
                    >= max_non_empty_rows
                ):
                    break

            inspections.append(
                WorksheetInspection(
                    title=(
                        worksheet.title
                    ),
                    max_row=(
                        worksheet.max_row
                    ),
                    max_column=(
                        worksheet.max_column
                    ),
                    non_empty_rows=tuple(
                        non_empty_rows
                    ),
                )
            )

        return WorkbookInspection(
            sheet_names=tuple(
                workbook.sheetnames
            ),
            worksheets=tuple(
                inspections
            ),
        )

    finally:
        workbook.close()