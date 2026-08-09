import argparse
import base64
import json
import re
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from time import sleep, strptime
from typing import Any
from urllib.parse import urljoin

import httpx
from openpyxl import load_workbook

IDX_ORIGIN = "https://www.idx.id"

BLOCK_ORIGIN = "https://block.idx.id"


HIGHLIGHT_PATH = (
    "/en/market-data/statistical-reports/"
    "digital-statistic/monthly/highlights/"
    "statistical-highlight"
)

DELISTED_PAGE_PATH = (
    "/en/market-data/statistical-reports/"
    "digital-statistic/monthly/"
    "corporate-action-of-listed-companies/"
    "delisted-company"
)

REPORT_ENDPOINT = (
    "/primary/DigitalStatistic/"
    "GetReportData"
)

LISTING_ACTIVITY_DELISTING_ENDPOINT = (
    "/primary/ListingActivity/"
    "GetDelisting"
)

LISTING_ACTIVITY_IPO_RELISTING_ENDPOINT = (
    "/primary/ListingActivity/"
    "GetIpoRelisting"
)

ANNUAL_REPORT_URL = (
    IDX_ORIGIN
    + "/en/about-idx/annual-report/"
)


DEFAULT_IPO_SNAPSHOT = (
    "data/derived/"
    "idx_ipo_history_snapshot.json"
)

DEFAULT_OUTPUT = (
    "data/derived/"
    "idx_lifecycle_evidence_batch.json"
)


DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 "
        "Indonesia-Market-Intelligence/"
        "lifecycle-evidence-batch"
    ),
    "Accept": "*/*",
}


PAGE_SIZE = 200

MAX_PAGES = 20


@dataclass(
    frozen=True,
    slots=True,
)
class DelistingRecord:
    symbol: str
    company_name: str | None
    listing_date: str | None
    delisting_date: str
    source: str
    year: int
    month: int


@dataclass(
    frozen=True,
    slots=True,
)
class HighlightEvidence:
    year: int
    month: int
    http_status: int | None
    parse_status: str
    issuer_count: int | None
    error: str | None


@dataclass(
    frozen=True,
    slots=True,
)
class DelistingMonthEvidence:
    year: int
    month: int
    highlight: HighlightEvidence
    html_http_status: int | None
    html_records: tuple[
        DelistingRecord,
        ...
    ]
    xlsx_http_status: int | None
    xlsx_records: tuple[
        DelistingRecord,
        ...
    ]
    resolved_records: tuple[
        DelistingRecord,
        ...
    ]
    resolution_status: str


@dataclass(
    frozen=True,
    slots=True,
)
class EndpointProbe:
    origin: str
    mode: str
    http_status: int | None
    content_type: str | None
    message: str | None
    error: str | None


@dataclass(
    frozen=True,
    slots=True,
)
class RelistingRecord:
    symbol: str
    company_name: str | None
    event_date: str
    year: int
    raw_status: str | None


@dataclass(
    frozen=True,
    slots=True,
)
class RelistingYearEvidence:
    year: int
    pages_requested: int
    total_rows: int
    records: tuple[
        RelistingRecord,
        ...
    ]
    response_coverage: bool
    error: str | None


@dataclass(
    frozen=True,
    slots=True,
)
class AnnualReportLink:
    label: str
    url: str


@dataclass(
    frozen=True,
    slots=True,
)
class LifecycleBatchSnapshot:
    generated_at: str
    snapshot_version: str
    ipo_snapshot_path: str
    ipo_snapshot_found: bool
    ipo_canonical_rows: int | None
    delisting_endpoint_probes: tuple[
        EndpointProbe,
        ...
    ]
    delisting_months: tuple[
        DelistingMonthEvidence,
        ...
    ]
    delisting_records: tuple[
        DelistingRecord,
        ...
    ]
    relisting_years: tuple[
        RelistingYearEvidence,
        ...
    ]
    relisting_records: tuple[
        RelistingRecord,
        ...
    ]
    annual_report_links: tuple[
        AnnualReportLink,
        ...
    ]
    delisting_positive_months: int
    delisting_resolved_positive_months: int
    delisting_unknown_months: int
    relisting_coverage_years: int
    lifecycle_event_ready: bool
    strict_pit_ready: bool


class VisibleTextParser(
    HTMLParser
):
    def __init__(
        self,
    ) -> None:
        super().__init__()

        self.parts: list[
            str
        ] = []

    def handle_data(
        self,
        data: str,
    ) -> None:
        normalized = " ".join(
            data.split()
        )

        if normalized:
            self.parts.append(
                normalized
            )

    def text(
        self,
    ) -> str:
        return " ".join(
            self.parts
        )


class TableParser(
    HTMLParser
):
    def __init__(
        self,
    ) -> None:
        super().__init__()

        self.rows: list[
            list[str]
        ] = []

        self._in_row = False

        self._cell_depth = 0

        self._current_row: list[
            str
        ] = []

        self._cell_parts: list[
            str
        ] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[
            tuple[
                str,
                str | None,
            ]
        ],
    ) -> None:
        del attrs

        normalized = (
            tag.casefold()
        )

        if normalized == "tr":
            self._in_row = True
            self._current_row = []

        elif (
            normalized
            in {
                "td",
                "th",
            }
            and self._in_row
        ):
            self._cell_depth += 1

            if self._cell_depth == 1:
                self._cell_parts = []

    def handle_data(
        self,
        data: str,
    ) -> None:
        if (
            self._in_row
            and self._cell_depth > 0
        ):
            self._cell_parts.append(
                data
            )

    def handle_endtag(
        self,
        tag: str,
    ) -> None:
        normalized = (
            tag.casefold()
        )

        if (
            normalized
            in {
                "td",
                "th",
            }
            and self._in_row
            and self._cell_depth > 0
        ):
            self._cell_depth -= 1

            if self._cell_depth == 0:
                value = " ".join(
                    " ".join(
                        self._cell_parts
                    ).split()
                )

                self._current_row.append(
                    value
                )

                self._cell_parts = []

        elif (
            normalized == "tr"
            and self._in_row
        ):
            if self._current_row:
                self.rows.append(
                    self._current_row
                )

            self._current_row = []
            self._cell_parts = []
            self._cell_depth = 0
            self._in_row = False


class AnnualReportParser(
    HTMLParser
):
    def __init__(
        self,
        *,
        base_url: str,
    ) -> None:
        super().__init__()

        self.base_url = base_url

        self.links: list[
            AnnualReportLink
        ] = []

        self._href: str | None = None

        self._text_parts: list[
            str
        ] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[
            tuple[
                str,
                str | None,
            ]
        ],
    ) -> None:
        if tag.casefold() != "a":
            return

        attributes = dict(
            attrs
        )

        href = attributes.get(
            "href"
        )

        if not href:
            return

        self._href = urljoin(
            self.base_url,
            href,
        )

        self._text_parts = []

    def handle_data(
        self,
        data: str,
    ) -> None:
        if self._href is None:
            return

        normalized = " ".join(
            data.split()
        )

        if normalized:
            self._text_parts.append(
                normalized
            )

    def handle_endtag(
        self,
        tag: str,
    ) -> None:
        if (
            tag.casefold() != "a"
            or self._href is None
        ):
            return

        label = " ".join(
            self._text_parts
        ).strip()

        lowered = (
            self._href.casefold()
        )

        if (
            ".pdf" in lowered
            or "/media/" in lowered
        ):
            self.links.append(
                AnnualReportLink(
                    label=label,
                    url=self._href,
                )
            )

        self._href = None

        self._text_parts = []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run a consolidated official "
            "IDX lifecycle evidence batch "
            "for delisting and relisting."
        )
    )

    parser.add_argument(
        "--start-year",
        type=int,
        default=2020,
    )

    parser.add_argument(
        "--end-year",
        type=int,
        default=2026,
    )

    parser.add_argument(
        "--end-month",
        type=int,
        default=8,
    )

    parser.add_argument(
        "--relisting-start-year",
        type=int,
        default=1990,
    )

    parser.add_argument(
        "--ipo-snapshot",
        default=(
            DEFAULT_IPO_SNAPSHOT
        ),
    )

    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
    )

    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
    )

    parser.add_argument(
        "--pause",
        type=float,
        default=0.10,
    )

    return parser.parse_args()


def validate_args(
    args: argparse.Namespace,
) -> None:
    if not (
        1900
        <= args.start_year
        <= args.end_year
        <= 2100
    ):
        raise ValueError(
            "Require 1900 <= start-year "
            "<= end-year <= 2100."
        )

    if not (
        1
        <= args.end_month
        <= 12
    ):
        raise ValueError(
            "end-month must be "
            "between 1 and 12."
        )

    if not (
        1900
        <= args.relisting_start_year
        <= args.end_year
    ):
        raise ValueError(
            "Invalid relisting-start-year."
        )

    if args.timeout <= 0:
        raise ValueError(
            "timeout must be positive."
        )

    if args.pause < 0:
        raise ValueError(
            "pause cannot be negative."
        )

    if not args.output.strip():
        raise ValueError(
            "output cannot be empty."
        )


def compact_text(
    value: str,
) -> str:
    return " ".join(
        value.split()
    )


def normalize_symbol(
    value: Any,
) -> str | None:
    if not isinstance(
        value,
        str,
    ):
        return None

    normalized = (
        value.strip()
        .upper()
    )

    return (
        normalized
        if normalized
        else None
    )


def normalize_optional_text(
    value: Any,
) -> str | None:
    if value is None:
        return None

    normalized = str(
        value
    ).strip()

    return (
        normalized
        if normalized
        else None
    )


def normalize_date(
    value: Any,
) -> str | None:
    if value is None:
        return None

    if isinstance(
        value,
        datetime,
    ):
        return (
            value.date()
            .isoformat()
        )

    if isinstance(
        value,
        date,
    ):
        return value.isoformat()

    if not isinstance(
        value,
        str,
    ):
        return None

    text = value.strip()

    if not text:
        return None

    if (
        len(text) >= 10
        and text[4] == "-"
        and text[7] == "-"
    ):
        try:
            parsed = date.fromisoformat(
                text[:10]
            )

        except ValueError:
            pass

        else:
            if parsed.year != 1:
                return parsed.isoformat()

    formats = (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%d-%b-%y",
    )

    for date_format in formats:
        try:
            parsed_time = strptime(
                text,
                date_format,
            )

        except ValueError:
            continue

        parsed = date(
            parsed_time.tm_year,
            parsed_time.tm_mon,
            parsed_time.tm_mday,
        )

        if parsed.year == 1:
            return None

        return parsed.isoformat()

    return None


def make_filter(
    *,
    year: int,
    month: int,
) -> str:
    payload = {
        "year": str(
            year
        ),
        "month": str(
            month
        ),
        "quarter": 0,
        "type": "monthly",
    }

    encoded = json.dumps(
        payload,
        separators=(
            ",",
            ":",
        ),
    ).encode(
        "utf-8"
    )

    return base64.b64encode(
        encoded
    ).decode(
        "ascii"
    )


def highlight_url(
    *,
    year: int,
    month: int,
) -> str:
    return (
        IDX_ORIGIN
        + HIGHLIGHT_PATH
        + "?filter="
        + make_filter(
            year=year,
            month=month,
        )
    )


def delisted_page_url(
    *,
    year: int,
    month: int,
) -> str:
    return (
        IDX_ORIGIN
        + DELISTED_PAGE_PATH
        + "?filter="
        + make_filter(
            year=year,
            month=month,
        )
    )


def parse_highlight_delisting(
    html: str,
) -> tuple[
    str,
    int | None,
]:
    parser = VisibleTextParser()

    parser.feed(
        html
    )

    parser.close()

    text = parser.text()

    marker = "Stock Delisting"

    position = text.casefold().find(
        marker.casefold()
    )

    if position < 0:
        return (
            "SECTION_NOT_FOUND",
            None,
        )

    section = text[
        position:
        position + 500
    ]

    rights_position = (
        section.casefold()
        .find(
            "rights announcement"
        )
    )

    if rights_position >= 0:
        section = section[
            :rights_position
        ]

    match = re.search(
        r"Issuers\s+([0-9][0-9,]*)",
        section,
        flags=re.IGNORECASE,
    )

    if match is not None:
        count = int(
            match.group(
                1
            ).replace(
                ",",
                "",
            )
        )

        return (
            "COUNT_PARSED",
            count,
        )

    if re.search(
        r"Issuers\s+Volume",
        section,
        flags=re.IGNORECASE,
    ):
        return (
            "EMPTY_FIELD",
            None,
        )

    return (
        "COUNT_NOT_PARSED",
        None,
    )


def fetch_highlight(
    *,
    client: httpx.Client,
    year: int,
    month: int,
) -> HighlightEvidence:
    url = highlight_url(
        year=year,
        month=month,
    )

    try:
        response = client.get(
            url
        )

    except httpx.HTTPError as exc:
        return HighlightEvidence(
            year=year,
            month=month,
            http_status=None,
            parse_status="HTTP_ERROR",
            issuer_count=None,
            error=(
                f"{type(exc).__name__}: "
                f"{exc}"
            ),
        )

    if response.status_code != 200:
        return HighlightEvidence(
            year=year,
            month=month,
            http_status=(
                response.status_code
            ),
            parse_status=(
                "HTTP_NON_200"
            ),
            issuer_count=None,
            error=(
                "Unexpected HTTP "
                f"{response.status_code}."
            ),
        )

    (
        parse_status,
        issuer_count,
    ) = parse_highlight_delisting(
        response.text
    )

    return HighlightEvidence(
        year=year,
        month=month,
        http_status=(
            response.status_code
        ),
        parse_status=(
            parse_status
        ),
        issuer_count=issuer_count,
        error=None,
    )


def normalize_header(
    value: Any,
) -> str:
    if value is None:
        return ""

    text = str(
        value
    ).casefold()

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return " ".join(
        text.split()
    )


def is_code_header(
    value: str,
) -> bool:
    return value in {
        "code",
        "kode",
        "kode emiten",
    }


def is_company_header(
    value: str,
) -> bool:
    return value in {
        "company name",
        "stock name",
        "nama perusahaan",
        "nama emiten",
    }


def is_listing_header(
    value: str,
) -> bool:
    return value in {
        "listing date",
        "tanggal pencatatan",
    }


def is_delisting_header(
    value: str,
) -> bool:
    return value in {
        "delisting date",
        "tanggal delisting",
    }


def header_positions(
    row: list[
        Any
    ],
) -> dict[
    str,
    int
] | None:
    positions: dict[
        str,
        int
    ] = {}

    for index, value in enumerate(
        row
    ):
        normalized = normalize_header(
            value
        )

        if is_code_header(
            normalized
        ):
            positions[
                "code"
            ] = index

        elif is_company_header(
            normalized
        ):
            positions[
                "company"
            ] = index

        elif is_listing_header(
            normalized
        ):
            positions[
                "listing"
            ] = index

        elif is_delisting_header(
            normalized
        ):
            positions[
                "delisting"
            ] = index

    if (
        "code" in positions
        and "delisting" in positions
    ):
        return positions

    return None


def safe_cell(
    row: list[
        Any
    ],
    index: int | None,
) -> Any:
    if index is None:
        return None

    if index >= len(
        row
    ):
        return None

    return row[
        index
    ]


def parse_table_rows(
    *,
    rows: list[
        list[Any]
    ],
    source: str,
    year: int,
    month: int,
) -> tuple[
    DelistingRecord,
    ...
]:
    header_index = None

    positions = None

    for index, row in enumerate(
        rows
    ):
        detected = header_positions(
            row
        )

        if detected is None:
            continue

        header_index = index
        positions = detected
        break

    if (
        header_index is None
        or positions is None
    ):
        return ()

    records = []

    for row in rows[
        header_index + 1:
    ]:
        symbol = normalize_symbol(
            safe_cell(
                row,
                positions.get(
                    "code"
                ),
            )
        )

        delisting_date = (
            normalize_date(
                safe_cell(
                    row,
                    positions.get(
                        "delisting"
                    ),
                )
            )
        )

        if (
            symbol is None
            or delisting_date is None
        ):
            continue

        company_name = (
            normalize_optional_text(
                safe_cell(
                    row,
                    positions.get(
                        "company"
                    ),
                )
            )
        )

        listing_date = (
            normalize_date(
                safe_cell(
                    row,
                    positions.get(
                        "listing"
                    ),
                )
            )
        )

        records.append(
            DelistingRecord(
                symbol=symbol,
                company_name=(
                    company_name
                ),
                listing_date=(
                    listing_date
                ),
                delisting_date=(
                    delisting_date
                ),
                source=source,
                year=year,
                month=month,
            )
        )

    return tuple(
        records
    )


def fetch_delisted_html(
    *,
    client: httpx.Client,
    year: int,
    month: int,
) -> tuple[
    int | None,
    tuple[
        DelistingRecord,
        ...
    ],
]:
    try:
        response = client.get(
            delisted_page_url(
                year=year,
                month=month,
            )
        )

    except httpx.HTTPError:
        return (
            None,
            (),
        )

    if response.status_code != 200:
        return (
            response.status_code,
            (),
        )

    parser = TableParser()

    parser.feed(
        response.text
    )

    parser.close()

    records = parse_table_rows(
        rows=[
            list(
                row
            )
            for row in parser.rows
        ],
        source="DIGITAL_STATISTIC_HTML",
        year=year,
        month=month,
    )

    return (
        response.status_code,
        records,
    )


def fetch_delisted_xlsx(
    *,
    client: httpx.Client,
    year: int,
    month: int,
) -> tuple[
    int | None,
    tuple[
        DelistingRecord,
        ...
    ],
]:
    params = {
        "type": "excel",
        "periodType": "monthly",
        "periodYear": year,
        "periodMonth": month,
        "cumulative": "false",
        "filecode": "SDelisting",
        "filename": (
            "Delisted Company"
        ),
    }

    try:
        response = client.get(
            IDX_ORIGIN
            + REPORT_ENDPOINT,
            params=params,
        )

    except httpx.HTTPError:
        return (
            None,
            (),
        )

    if response.status_code != 200:
        return (
            response.status_code,
            (),
        )

    if not response.content.startswith(
        b"PK\x03\x04"
    ):
        return (
            response.status_code,
            (),
        )

    try:
        workbook = load_workbook(
            filename=BytesIO(
                response.content
            ),
            read_only=True,
            data_only=True,
        )

    except (
        OSError,
        ValueError,
    ):
        return (
            response.status_code,
            (),
        )

    all_records = []

    try:
        for worksheet in (
            workbook.worksheets
        ):
            rows = [
                list(
                    row
                )
                for row in (
                    worksheet
                    .iter_rows(
                        values_only=True
                    )
                )
            ]

            records = parse_table_rows(
                rows=rows,
                source=(
                    "DIGITAL_STATISTIC_XLSX"
                ),
                year=year,
                month=month,
            )

            all_records.extend(
                records
            )

    finally:
        workbook.close()

    return (
        response.status_code,
        tuple(
            all_records
        ),
    )


def record_key(
    record: DelistingRecord,
) -> tuple[
    str,
    str,
]:
    return (
        record.symbol,
        record.delisting_date,
    )


def reconcile_delisting_records(
    *,
    html_records: tuple[
        DelistingRecord,
        ...
    ],
    xlsx_records: tuple[
        DelistingRecord,
        ...
    ],
) -> tuple[
    DelistingRecord,
    ...
]:
    chosen: dict[
        tuple[
            str,
            str,
        ],
        DelistingRecord,
    ] = {}

    for record in html_records:
        chosen[
            record_key(
                record
            )
        ] = record

    for record in xlsx_records:
        chosen[
            record_key(
                record
            )
        ] = record

    return tuple(
        sorted(
            chosen.values(),
            key=lambda record: (
                record.delisting_date,
                record.symbol,
            ),
        )
    )


def build_month_evidence(
    *,
    client: httpx.Client,
    highlight: HighlightEvidence,
) -> DelistingMonthEvidence:
    should_fetch_detail = (
        highlight.issuer_count
        is not None
        and highlight.issuer_count
        > 0
    )

    if not should_fetch_detail:
        return DelistingMonthEvidence(
            year=highlight.year,
            month=highlight.month,
            highlight=highlight,
            html_http_status=None,
            html_records=(),
            xlsx_http_status=None,
            xlsx_records=(),
            resolved_records=(),
            resolution_status=(
                "NO_POSITIVE_COUNT"
                if (
                    highlight.issuer_count
                    == 0
                )
                else "COUNT_UNKNOWN"
            ),
        )

    (
        html_status,
        html_records,
    ) = fetch_delisted_html(
        client=client,
        year=highlight.year,
        month=highlight.month,
    )

    (
        xlsx_status,
        xlsx_records,
    ) = fetch_delisted_xlsx(
        client=client,
        year=highlight.year,
        month=highlight.month,
    )

    resolved = (
        reconcile_delisting_records(
            html_records=html_records,
            xlsx_records=xlsx_records,
        )
    )

    if (
        highlight.issuer_count
        == len(
            resolved
        )
    ):
        status = (
            "RESOLVED"
        )

    elif resolved:
        status = (
            "PARTIAL"
        )

    else:
        status = (
            "POSITIVE_COUNT_NO_ROWS"
        )

    return DelistingMonthEvidence(
        year=highlight.year,
        month=highlight.month,
        highlight=highlight,
        html_http_status=html_status,
        html_records=html_records,
        xlsx_http_status=xlsx_status,
        xlsx_records=xlsx_records,
        resolved_records=resolved,
        resolution_status=status,
    )


def months_to_scan(
    *,
    start_year: int,
    end_year: int,
    end_month: int,
) -> tuple[
    tuple[
        int,
        int,
    ],
    ...
]:
    values = []

    for year in range(
        start_year,
        end_year + 1,
    ):
        max_month = (
            end_month
            if year == end_year
            else 12
        )

        for month in range(
            1,
            max_month + 1,
        ):
            values.append(
                (
                    year,
                    month,
                )
            )

    return tuple(
        values
    )


def probe_delisting_endpoint(
    *,
    client: httpx.Client,
    origin: str,
    with_status: bool,
) -> EndpointProbe:
    params = (
        {
            "Status": "delisting",
        }
        if with_status
        else None
    )

    mode = (
        "STATUS_DELISTING"
        if with_status
        else "BARE"
    )

    try:
        response = client.get(
            origin
            + LISTING_ACTIVITY_DELISTING_ENDPOINT,
            params=params,
        )

    except httpx.HTTPError as exc:
        return EndpointProbe(
            origin=origin,
            mode=mode,
            http_status=None,
            content_type=None,
            message=None,
            error=(
                f"{type(exc).__name__}: "
                f"{exc}"
            ),
        )

    content_type = (
        response.headers.get(
            "content-type"
        )
    )

    message = None

    try:
        payload = response.json()

    except ValueError:
        body = compact_text(
            response.text
        )

        message = (
            body[:1000]
            if body
            else None
        )

    else:
        if isinstance(
            payload,
            dict,
        ):
            candidate = payload.get(
                "Message"
            )

            if candidate is not None:
                message = str(
                    candidate
                )

            else:
                message = (
                    json.dumps(
                        payload,
                        ensure_ascii=False,
                    )[:1000]
                )

        else:
            message = str(
                payload
            )[:1000]

    return EndpointProbe(
        origin=origin,
        mode=mode,
        http_status=(
            response.status_code
        ),
        content_type=content_type,
        message=message,
        error=None,
    )


def parse_relisting_row(
    *,
    row: dict[
        str,
        Any,
    ],
    year: int,
) -> RelistingRecord | None:
    symbol = normalize_symbol(
        row.get(
            "KodeEmiten"
        )
    )

    event_date = normalize_date(
        row.get(
            "TanggalPencatatan"
        )
    )

    if (
        symbol is None
        or event_date is None
    ):
        return None

    return RelistingRecord(
        symbol=symbol,
        company_name=(
            normalize_optional_text(
                row.get(
                    "NamaEmiten"
                )
            )
        ),
        event_date=event_date,
        year=year,
        raw_status=(
            normalize_optional_text(
                row.get(
                    "RencanaStatus"
                )
            )
        ),
    )


def fetch_relisting_year(
    *,
    client: httpx.Client,
    year: int,
    pause: float,
) -> RelistingYearEvidence:
    records = []

    pages_requested = 0

    coverage = False

    error = None

    for page in range(
        1,
        MAX_PAGES + 1,
    ):
        params = {
            "Status": "relisting",
            "Year": year,
            "indexfrom": page,
            "pagesize": PAGE_SIZE,
        }

        try:
            response = client.get(
                (
                    IDX_ORIGIN
                    + LISTING_ACTIVITY_IPO_RELISTING_ENDPOINT
                ),
                params=params,
            )

        except httpx.HTTPError as exc:
            error = (
                f"{type(exc).__name__}: "
                f"{exc}"
            )
            break

        pages_requested += 1

        if response.status_code != 200:
            error = (
                "Unexpected HTTP "
                f"{response.status_code}."
            )
            break

        try:
            payload = response.json()

        except ValueError:
            error = (
                "Response is not JSON."
            )
            break

        if not isinstance(
            payload,
            dict,
        ):
            error = (
                "JSON root is not "
                "an object."
            )
            break

        criteria = payload.get(
            "SearchCriteria"
        )

        rows = payload.get(
            "Result"
        )

        if not isinstance(
            criteria,
            dict,
        ):
            error = (
                "SearchCriteria missing."
            )
            break

        if not isinstance(
            rows,
            list,
        ):
            error = (
                "Result is not an array."
            )
            break

        if str(
            criteria.get(
                "Year"
            )
        ) != str(
            year
        ):
            error = (
                "Year echo mismatch."
            )
            break

        if (
            criteria.get(
                "Status"
            )
            != "relisting"
        ):
            error = (
                "Status echo mismatch."
            )
            break

        try:
            echoed_page = int(
                criteria.get(
                    "indexfrom"
                )
            )

            echoed_pagesize = int(
                criteria.get(
                    "pagesize"
                )
            )

        except (
            TypeError,
            ValueError,
        ):
            error = (
                "Pagination echo invalid."
            )
            break

        if echoed_page != page:
            error = (
                "Page echo mismatch."
            )
            break

        if echoed_pagesize != PAGE_SIZE:
            error = (
                "Page size echo mismatch."
            )
            break

        for row in rows:
            if not isinstance(
                row,
                dict,
            ):
                continue

            record = parse_relisting_row(
                row=row,
                year=year,
            )

            if record is not None:
                records.append(
                    record
                )

        if len(
            rows
        ) < PAGE_SIZE:
            coverage = True
            break

        if pause > 0:
            sleep(
                pause
            )

    return RelistingYearEvidence(
        year=year,
        pages_requested=(
            pages_requested
        ),
        total_rows=len(
            records
        ),
        records=tuple(
            records
        ),
        response_coverage=(
            coverage
            and error is None
        ),
        error=error,
    )


def fetch_annual_report_links(
    *,
    client: httpx.Client,
) -> tuple[
    AnnualReportLink,
    ...
]:
    try:
        response = client.get(
            ANNUAL_REPORT_URL
        )

    except httpx.HTTPError:
        return ()

    if response.status_code != 200:
        return ()

    parser = AnnualReportParser(
        base_url=str(
            response.url
        )
    )

    parser.feed(
        response.text
    )

    parser.close()

    unique: dict[
        str,
        AnnualReportLink,
    ] = {}

    for link in parser.links:
        unique[
            link.url
        ] = link

    return tuple(
        unique.values()
    )


def load_ipo_summary(
    path: Path,
) -> tuple[
    bool,
    int | None,
]:
    if not path.exists():
        return (
            False,
            None,
        )

    try:
        payload = json.loads(
            path.read_text(
                encoding="utf-8"
            )
        )

    except (
        OSError,
        json.JSONDecodeError,
    ):
        return (
            False,
            None,
        )

    value = payload.get(
        "total_canonical_rows"
    )

    if isinstance(
        value,
        int,
    ):
        return (
            True,
            value,
        )

    return (
        True,
        None,
    )


def write_snapshot(
    *,
    snapshot: LifecycleBatchSnapshot,
    path: Path,
) -> None:
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    path.write_text(
        json.dumps(
            asdict(
                snapshot
            ),
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> None:
    args = parse_args()

    validate_args(
        args
    )

    ipo_path = Path(
        args.ipo_snapshot
    )

    output_path = Path(
        args.output
    )

    (
        ipo_found,
        ipo_rows,
    ) = load_ipo_summary(
        ipo_path
    )

    print(
        "Indonesia Market Intelligence"
    )

    print(
        "IDX Lifecycle Evidence Batch V1"
    )

    print(
        "--------------------------------"
    )

    print(
        f"IPO snapshot       : "
        f"{ipo_path}"
    )

    print(
        f"IPO snapshot found : "
        f"{ipo_found}"
    )

    print(
        f"IPO canonical rows : "
        f"{ipo_rows}"
    )

    print()

    endpoint_probes = []

    delisting_months = []

    relisting_years = []

    with httpx.Client(
        timeout=args.timeout,
        follow_redirects=True,
        headers=DEFAULT_HEADERS,
    ) as client:
        print(
            "1. DELISTING ENDPOINT PROBES"
        )

        for origin in (
            IDX_ORIGIN,
            BLOCK_ORIGIN,
        ):
            for with_status in (
                False,
                True,
            ):
                probe = (
                    probe_delisting_endpoint(
                        client=client,
                        origin=origin,
                        with_status=(
                            with_status
                        ),
                    )
                )

                endpoint_probes.append(
                    probe
                )

                print(
                    f"  {origin} "
                    f"{probe.mode:<16} "
                    f"HTTP="
                    f"{probe.http_status}"
                )

                if args.pause > 0:
                    sleep(
                        args.pause
                    )

        print()

        print(
            "2. MONTHLY DELISTING EVIDENCE"
        )

        scan_months = months_to_scan(
            start_year=(
                args.start_year
            ),
            end_year=(
                args.end_year
            ),
            end_month=(
                args.end_month
            ),
        )

        for index, (
            year,
            month,
        ) in enumerate(
            scan_months,
            start=1,
        ):
            highlight = fetch_highlight(
                client=client,
                year=year,
                month=month,
            )

            evidence = (
                build_month_evidence(
                    client=client,
                    highlight=highlight,
                )
            )

            delisting_months.append(
                evidence
            )

            if (
                highlight.issuer_count
                is not None
                and highlight.issuer_count
                > 0
            ):
                print(
                    f"  {year}-{month:02d} "
                    f"count="
                    f"{highlight.issuer_count} "
                    f"records="
                    f"{len(evidence.resolved_records)} "
                    f"{evidence.resolution_status}"
                )

            elif (
                highlight.parse_status
                not in {
                    "COUNT_PARSED",
                    "EMPTY_FIELD",
                }
            ):
                print(
                    f"  {year}-{month:02d} "
                    f"{highlight.parse_status}"
                )

            if (
                index
                < len(
                    scan_months
                )
                and args.pause > 0
            ):
                sleep(
                    args.pause
                )

        print()

        print(
            "3. RELISTING COVERAGE"
        )

        for year in range(
            args.relisting_start_year,
            args.end_year + 1,
        ):
            result = fetch_relisting_year(
                client=client,
                year=year,
                pause=args.pause,
            )

            relisting_years.append(
                result
            )

            if (
                result.total_rows > 0
                or result.error is not None
            ):
                print(
                    f"  {year} "
                    f"rows="
                    f"{result.total_rows} "
                    f"coverage="
                    f"{result.response_coverage} "
                    f"error="
                    f"{result.error}"
                )

            if args.pause > 0:
                sleep(
                    args.pause
                )

        print()

        print(
            "4. ANNUAL REPORT INVENTORY"
        )

        annual_links = (
            fetch_annual_report_links(
                client=client
            )
        )

        print(
            f"  PDF/report links : "
            f"{len(annual_links)}"
        )

    delisting_records_map: dict[
        tuple[
            str,
            str,
        ],
        DelistingRecord,
    ] = {}

    for evidence in (
        delisting_months
    ):
        for record in (
            evidence
            .resolved_records
        ):
            delisting_records_map[
                (
                    record.symbol,
                    record.delisting_date,
                )
            ] = record

    delisting_records = tuple(
        sorted(
            delisting_records_map.values(),
            key=lambda record: (
                record.delisting_date,
                record.symbol,
            ),
        )
    )

    relisting_records_map: dict[
        tuple[
            str,
            str,
        ],
        RelistingRecord,
    ] = {}

    for evidence in (
        relisting_years
    ):
        for record in (
            evidence.records
        ):
            relisting_records_map[
                (
                    record.symbol,
                    record.event_date,
                )
            ] = record

    relisting_records = tuple(
        sorted(
            relisting_records_map.values(),
            key=lambda record: (
                record.event_date,
                record.symbol,
            ),
        )
    )

    positive_months = sum(
        (
            evidence
            .highlight
            .issuer_count
            or 0
        )
        > 0
        for evidence in delisting_months
    )

    resolved_positive_months = sum(
        (
            evidence
            .highlight
            .issuer_count
            is not None
            and evidence
            .highlight
            .issuer_count
            > 0
            and evidence
            .resolution_status
            == "RESOLVED"
        )
        for evidence in delisting_months
    )

    unknown_months = sum(
        evidence.highlight.issuer_count
        is None
        for evidence in delisting_months
    )

    relisting_coverage_years = sum(
        result.response_coverage
        for result in relisting_years
    )

    total_relisting_years = len(
        relisting_years
    )

    delisting_event_ready = (
        unknown_months == 0
        and positive_months
        == resolved_positive_months
    )

    relisting_ready = (
        relisting_coverage_years
        == total_relisting_years
    )

    lifecycle_event_ready = (
        ipo_found
        and delisting_event_ready
        and relisting_ready
    )

    snapshot = (
        LifecycleBatchSnapshot(
            generated_at=(
                datetime.now(
                    UTC
                ).isoformat()
            ),
            snapshot_version=(
                "idx_lifecycle_evidence_v1"
            ),
            ipo_snapshot_path=str(
                ipo_path
            ),
            ipo_snapshot_found=(
                ipo_found
            ),
            ipo_canonical_rows=(
                ipo_rows
            ),
            delisting_endpoint_probes=tuple(
                endpoint_probes
            ),
            delisting_months=tuple(
                delisting_months
            ),
            delisting_records=(
                delisting_records
            ),
            relisting_years=tuple(
                relisting_years
            ),
            relisting_records=(
                relisting_records
            ),
            annual_report_links=(
                annual_links
            ),
            delisting_positive_months=(
                positive_months
            ),
            delisting_resolved_positive_months=(
                resolved_positive_months
            ),
            delisting_unknown_months=(
                unknown_months
            ),
            relisting_coverage_years=(
                relisting_coverage_years
            ),
            lifecycle_event_ready=(
                lifecycle_event_ready
            ),
            strict_pit_ready=False,
        )
    )

    write_snapshot(
        snapshot=snapshot,
        path=output_path,
    )

    print()

    print(
        "SUMMARY"
    )

    print(
        f"Delisting months scanned : "
        f"{len(delisting_months)}"
    )

    print(
        f"Positive months          : "
        f"{positive_months}"
    )

    print(
        f"Positive resolved        : "
        f"{resolved_positive_months}"
    )

    print(
        f"Unknown count months     : "
        f"{unknown_months}"
    )

    print(
        f"Delisting records        : "
        f"{len(delisting_records)}"
    )

    print(
        f"Relisting years scanned  : "
        f"{total_relisting_years}"
    )

    print(
        f"Relisting coverage       : "
        f"{relisting_coverage_years}/"
        f"{total_relisting_years}"
    )

    print(
        f"Relisting records        : "
        f"{len(relisting_records)}"
    )

    print(
        f"Annual report links      : "
        f"{len(annual_links)}"
    )

    print(
        f"Snapshot                 : "
        f"{output_path}"
    )

    print()

    print(
        "READINESS"
    )

    print(
        "IPO evidence available : "
        + (
            "YES"
            if ipo_found
            else "NO"
        )
    )

    print(
        "Delisting event ready  : "
        + (
            "YES"
            if delisting_event_ready
            else "NO"
        )
    )

    print(
        "Relisting query ready  : "
        + (
            "YES"
            if relisting_ready
            else "NO"
        )
    )

    print(
        "Lifecycle event ready  : "
        + (
            "YES"
            if lifecycle_event_ready
            else "NO"
        )
    )

    print(
        "Strict PIT ready       : NO"
    )

    print()

    print(
        "IMPORTANT:"
    )

    print(
        "EMPTY_FIELD is not promoted "
        "to a confirmed zero delisting "
        "count."
    )

    print(
        "Partial Delisting is excluded "
        "from full-company lifecycle "
        "delisting events."
    )

    print(
        "Historical available_at remains "
        "UNKNOWN."
    )

    print()

    print(
        "DATABASE WRITE:"
    )

    print(
        "ENABLED : NO"
    )


if __name__ == "__main__":
    main()